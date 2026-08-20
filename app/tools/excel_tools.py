import json
import os
import re
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.adk.tools import ToolContext

MANDATORY_IDENTIFIERS = ["gtin", "upc", "ean", "mpn", "vpn", "asin"]


def _clean_json_string(json_str: str) -> str:
    """Extracts raw JSON content if wrapped in markdown code blocks."""
    match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", json_str, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return json_str.strip()


def validate_and_generate_excel(
    product_data_json: str,
    output_filename: str = "product_spec.xlsx",
    tool_context: Optional[ToolContext] = None,
) -> Dict[str, Any]:
    """Validates multi-SKU product specification JSON data and generates a styled multi-line Excel (.xlsx) file.

    Args:
        product_data_json: JSON string containing product family, SKU tree, and metadata.
        output_filename: Name of the output Excel file.

    Returns:
        A dictionary with validation status, missing fields, verified SKUs, and output file path.
    """
    cleaned_json = _clean_json_string(product_data_json)
    try:
        data = json.loads(cleaned_json)
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to parse product JSON: {str(e)}",
            "validation_passed": False,
        }

    # Extract top-level family & metadata
    product_family = data.get("product_family", data.get("product_info", {}).get("title", data.get("title", "Product Family")))
    brand = data.get("brand", data.get("product_info", {}).get("brand", "N/A"))
    category = data.get("category", data.get("product_info", {}).get("category", "N/A"))
    metadata = data.get("metadata", {})
    confidence_score = metadata.get("confidence_score", "N/A")
    timestamp = metadata.get("timestamp", "N/A")
    data_sources = metadata.get("data_sources", [])
    data_sources_str = ", ".join(data_sources) if isinstance(data_sources, list) else str(data_sources)

    # Extract SKUs list (handle multi-SKU array or single product fallback)
    raw_skus = data.get("skus")
    if not raw_skus or not isinstance(raw_skus, list):
        # Fallback to single product JSON format
        product_info = data.get("product_info", data)
        single_sku = {
            "sku_id": product_info.get("sku_id", "DEFAULT-SKU"),
            "variant_name": product_info.get("variant_name", product_info.get("title", "Standard Variant")),
            "price": product_info.get("price", "N/A"),
            "currency": product_info.get("currency", "USD"),
            "universal_identifiers": data.get("universal_identifiers", data.get("identifiers", {})),
            "attributes": data.get("specifications", data.get("attributes", {})),
        }
        raw_skus = [single_sku]

    # Validate mandatory IDs across all SKUs
    verified_skus = []
    missing_critical = []
    sku_rows = []

    for sku in raw_skus:
        sku_id = sku.get("sku_id", "N/A")
        variant_name = sku.get("variant_name", sku.get("title", "N/A"))
        price = f"{sku.get('price', 'N/A')} {sku.get('currency', '')}".strip()
        ids = sku.get("universal_identifiers", sku.get("identifiers", {}))
        attributes = sku.get("attributes", sku.get("specifications", {}))

        found_ids = {k.lower(): v for k, v in ids.items() if v and str(v).strip()}
        has_id = any(k in found_ids for k in MANDATORY_IDENTIFIERS)

        if not has_id:
            missing_critical.append(f"SKU {sku_id} missing mandatory universal identifiers (GTIN, UPC, EAN, MPN, VPN, or ASIN)")

        gtin = ids.get("gtin") or ids.get("gtin14") or "N/A"
        upc = ids.get("upc") or "N/A"
        ean = ids.get("ean") or "N/A"
        mpn = ids.get("mpn") or "N/A"
        vpn = ids.get("vpn") or "N/A"
        asin = ids.get("asin") or "N/A"
        unspsc = ids.get("unspsc") or "N/A"
        hs_code = ids.get("hs_code") or "N/A"

        attr_str = "; ".join([f"{k}: {v}" for k, v in attributes.items()]) if isinstance(attributes, dict) else str(attributes)

        sku_rows.append({
            "sku_id": sku_id,
            "variant_name": variant_name,
            "price": price,
            "gtin": gtin,
            "upc": upc,
            "ean": ean,
            "mpn": mpn,
            "vpn": vpn,
            "asin": asin,
            "unspsc": unspsc,
            "hs_code": hs_code,
            "attributes": attr_str,
            "status": "VERIFIED" if has_id else "MISSING IDs",
        })
        verified_skus.append(sku_id)

    validation_passed = len(missing_critical) == 0

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-SKU Catalog & IDs"

    # Styling Palette
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sub_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    navy_bold = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    regular_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title Block
    ws["A1"] = "UNIVERSAL PRODUCT INTELLIGENCE - MULTI-SKU SPECIFICATION SHEET"
    ws["A1"].font = title_font

    # Overview Metadata Block
    meta_info = [
        ("Product Family:", product_family),
        ("Brand:", brand),
        ("Category:", category),
        ("Confidence Score:", f"{confidence_score}"),
        ("Timestamp:", timestamp),
        ("Data Sources:", data_sources_str),
    ]

    for idx, (label, val) in enumerate(meta_info, start=3):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = str(val)
        ws[f"A{idx}"].font = navy_bold
        ws[f"B{idx}"].font = regular_font

    # Multi-Line SKU Table Headers
    headers = [
        "SKU ID", "Variant Name", "Price", "GTIN-14", "UPC-12", "EAN-13",
        "MPN", "VPN", "ASIN", "UNSPSC", "HS Code", "Attributes & Specifications", "Status"
    ]

    start_row = 10
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = h_text
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate SKU Rows
    for row_offset, row_data in enumerate(sku_rows, start=start_row + 1):
        row_vals = [
            row_data["sku_id"], row_data["variant_name"], row_data["price"],
            row_data["gtin"], row_data["upc"], row_data["ean"], row_data["mpn"],
            row_data["vpn"], row_data["asin"], row_data["unspsc"], row_data["hs_code"],
            row_data["attributes"], row_data["status"]
        ]
        for col_num, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_offset, column=col_num)
            cell.value = str(val)
            cell.font = regular_font
            cell.border = thin_border
            if col_num == 13: # Status column
                cell.font = Font(name="Calibri", size=11, bold=True, color="008000" if val == "VERIFIED" else "C00000")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # Save output file
    output_dir = os.path.join(os.getcwd(), "output")
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, output_filename)
    wb.save(file_path)

    return {
        "status": "success",
        "validation_passed": validation_passed,
        "missing_critical_fields": missing_critical,
        "verified_skus_count": len(verified_skus),
        "verified_skus": verified_skus,
        "metadata": {
            "confidence_score": confidence_score,
            "timestamp": timestamp,
            "data_sources": data_sources_str,
        },
        "output_file_path": file_path,
        "message": f"Multi-line Excel spreadsheet successfully created for {len(verified_skus)} SKUs at {file_path}",
    }
