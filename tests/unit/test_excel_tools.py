import json
import os
import openpyxl
import pytest
from app.tools.excel_tools import _clean_json_string, validate_and_generate_excel


def test_clean_json_string():
    raw_json = '{"key": "value"}'
    markdown_json = '```json\n{"key": "value"}\n```'
    markdown_plain = '```\n{"key": "value"}\n```'
    
    assert _clean_json_string(raw_json) == raw_json
    assert _clean_json_string(markdown_json) == raw_json
    assert _clean_json_string(markdown_plain) == raw_json


def test_validate_and_generate_excel_multi_sku(tmp_path):
    multi_sku_data = {
        "product_family": "CHANEL COCO MADEMOISELLE CRUSH ABSOLU",
        "brand": "CHANEL",
        "category": "Beauty & Personal Care > Fragrances",
        "skus": [
            {
                "sku_id": "CHA-50ML",
                "variant_name": "50ml Spray",
                "price": "125.00",
                "currency": "USD",
                "universal_identifiers": {
                    "gtin": "03145891165104",
                    "upc": "3145891165104",
                    "ean": "3145891165104",
                    "mpn": "116510",
                    "vpn": "CHA-116510",
                    "asin": "B00116510Y"
                },
                "attributes": {
                    "volume": "50ml"
                }
            },
            {
                "sku_id": "CHA-100ML",
                "variant_name": "100ml Spray",
                "price": "165.00",
                "currency": "USD",
                "universal_identifiers": {
                    "gtin": "03145891165203",
                    "upc": "3145891165203",
                    "ean": "3145891165203",
                    "mpn": "116520",
                    "vpn": "CHA-116520",
                    "asin": "B00116520X"
                },
                "attributes": {
                    "volume": "100ml"
                }
            }
        ],
        "metadata": {
            "confidence_score": 0.98,
            "timestamp": "2026-08-20T23:38:00Z",
            "data_sources": ["CHANEL Catalog", "Barcode Registry"]
        }
    }

    out_file = "multi_sku_test.xlsx"
    res = validate_and_generate_excel(json.dumps(multi_sku_data), output_filename=out_file)

    assert res["status"] == "success"
    assert res["validation_passed"] is True
    assert res["verified_skus_count"] == 2
    assert "CHA-50ML" in res["verified_skus"]
    assert "CHA-100ML" in res["verified_skus"]
    assert res["metadata"]["confidence_score"] == 0.98
    assert os.path.exists(res["output_file_path"])

    wb = openpyxl.load_workbook(res["output_file_path"])
    ws = wb["Multi-SKU Catalog & IDs"]
    assert ws["B3"].value == "CHANEL COCO MADEMOISELLE CRUSH ABSOLU"
    assert ws["B4"].value == "CHANEL"
    assert ws["A11"].value == "CHA-50ML"
    assert ws["A12"].value == "CHA-100ML"


def test_validate_and_generate_excel_single_fallback():
    sample_data = {
        "product_info": {
            "title": "Test Smartphone Ultra",
            "brand": "TechCorp",
            "category": "Mobile Devices",
            "price": "999.99",
            "currency": "USD"
        },
        "universal_identifiers": {
            "gtin": "01234567890123",
            "upc": "123456789012",
            "mpn": "TC-ULTRA-01"
        },
        "metadata": {
            "confidence_score": 0.95
        }
    }

    res = validate_and_generate_excel(json.dumps(sample_data))

    assert res["status"] == "success"
    assert res["validation_passed"] is True
    assert res["verified_skus_count"] == 1


def test_validate_and_generate_excel_missing_ids():
    sample_data = {
        "product_family": "Incomplete Product",
        "skus": [
            {
                "sku_id": "NO-ID-SKU",
                "universal_identifiers": {}
            }
        ]
    }

    res = validate_and_generate_excel(json.dumps(sample_data))

    assert res["status"] == "success"
    assert res["validation_passed"] is False
    assert len(res["missing_critical_fields"]) > 0


def test_validate_and_generate_excel_invalid_json():
    invalid_json_str = "{invalid json content"
    res = validate_and_generate_excel(invalid_json_str)

    assert res["status"] == "error"
    assert res["validation_passed"] is False
    assert "Failed to parse" in res["message"]
